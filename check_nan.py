#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
엑셀 파일의 NaN 값 확인 스크립트
"""

from openpyxl import load_workbook

def check_excel_for_nan(filename):
    """엑셀 파일에서 빈 셀 및 NaN 확인"""
    print(f"\n{'='*80}")
    print(f"파일 분석: {filename}")
    print(f"{'='*80}")
    
    try:
        wb = load_workbook(filename, data_only=True)
        sheet = wb.active
        
        # 헤더 읽기
        headers = []
        for col in range(1, sheet.max_column + 1):
            headers.append(sheet.cell(1, col).value)
        
        print(f"\n헤더: {headers}")
        
        # 빈 셀 찾기
        empty_rows = []
        for row_idx in range(2, min(sheet.max_row + 1, 20)):  # 처음 20행만 확인
            row_data = []
            has_empty = False
            for col_idx in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row_idx, col_idx).value
                row_data.append(cell_value)
                if cell_value is None or str(cell_value).strip() == '':
                    has_empty = True
            
            if has_empty:
                empty_rows.append((row_idx, row_data))
        
        if empty_rows:
            print(f"\n빈 셀이 있는 행 ({len(empty_rows)}개 발견):")
            for row_idx, row_data in empty_rows[:5]:  # 처음 5개만 표시
                print(f"  행 {row_idx}: {row_data}")
        else:
            print("\n✓ 처음 20행에서 빈 셀 없음")
        
        # 날짜 관련 열 확인
        date_col_indices = []
        for idx, header in enumerate(headers):
            if header and ('날짜' in str(header) or '기간' in str(header) or '출장' in str(header)):
                date_col_indices.append((idx + 1, header))
                print(f"\n날짜 관련 열 발견: 열 {idx + 1} - '{header}'")
                
                # 해당 열의 샘플 데이터 (빈 셀 포함)
                print(f"  샘플 데이터:")
                for row_idx in range(2, min(sheet.max_row + 1, 12)):
                    cell_value = sheet.cell(row_idx, idx + 1).value
                    if cell_value is None:
                        print(f"    행 {row_idx}: [빈 셀]")
                    else:
                        print(f"    행 {row_idx}: {cell_value}")
        
    except Exception as e:
        print(f"❌ 오류: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    files = ['근무상황목록.xlsx', '출장 목록 그리드.xlsx']
    
    for filename in files:
        check_excel_for_nan(filename)
