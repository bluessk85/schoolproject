#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
간단한 엑셀 파일 분석 도구 - openpyxl만 사용
"""

from openpyxl import load_workbook
import sys
from datetime import datetime

def analyze_excel_simple(filename):
    """openpyxl을 사용하여 엑셀 파일 분석"""
    print(f"\n{'='*80}")
    print(f"파일 분석: {filename}")
    print(f"{'='*80}")
    
    try:
        wb = load_workbook(filename, data_only=True)
        sheet = wb.active
        
        print(f"\n시트명: {sheet.title}")
        print(f"최대 행: {sheet.max_row}")
        print(f"최대 열: {sheet.max_column}")
        
        # 헤더 읽기 (첫 번째 행)
        headers = []
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, col).value
            headers.append(cell_value)
        
        print(f"\n열 헤더:")
        for i, header in enumerate(headers, 1):
            print(f"  {i}. '{header}'")
        
        # 처음 5행의 데이터 읽기
        print(f"\n처음 5행 데이터:")
        for row in range(1, min(6, sheet.max_row + 1)):
            row_data = []
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row, col)
                value = cell.value
                # 타입 정보 포함
                type_name = type(value).__name__
                row_data.append(f"{value} ({type_name})")
            print(f"  행 {row}: {row_data}")
        
        # 날짜 관련 컬럼 찾기
        print(f"\n날짜 관련 열 탐지:")
        date_keywords = ['날짜', 'date', '일자', '기간', '출장', '연가', '휴가', '조퇴']
        date_related_cols = []
        
        for i, header in enumerate(headers, 1):
            if header:
                header_lower = str(header).lower()
                if any(keyword in header_lower for keyword in date_keywords):
                    date_related_cols.append((i, header))
                    print(f"  ✓ 열 {i}: '{header}' 발견")
                    # 해당 열의 샘플 데이터
                    print(f"    샘플 데이터:")
                    for row in range(2, min(7, sheet.max_row + 1)):
                        cell = sheet.cell(row, i)
                        value = cell.value
                        type_name = type(value).__name__
                        print(f"      행 {row}: {value} (타입: {type_name})")
        
        if not date_related_cols:
            print("  ✗ 날짜 관련 키워드가 포함된 열을 찾지 못했습니다.")
            print("\n  모든 열의 데이터 타입 확인:")
            for i, header in enumerate(headers, 1):
                print(f"\n  열 {i}: '{header}'")
                for row in range(2, min(7, sheet.max_row + 1)):
                    cell = sheet.cell(row, i)
                    value = cell.value
                    type_name = type(value).__name__
                    print(f"    행 {row}: {value} (타입: {type_name})")
        
        return date_related_cols
        
    except Exception as e:
        print(f"❌ 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        return []

if __name__ == "__main__":
    files = ['근무상황목록.xlsx', '출장 목록 그리드.xlsx']
    
    all_date_cols = {}
    for filename in files:
        date_cols = analyze_excel_simple(filename)
        all_date_cols[filename] = date_cols
    
    print(f"\n{'='*80}")
    print("분석 요약")
    print(f"{'='*80}")
    for filename, cols in all_date_cols.items():
        print(f"\n{filename}:")
        if cols:
            for col_idx, col_name in cols:
                print(f"  - 열 {col_idx}: '{col_name}'")
        else:
            print("  - 날짜 관련 열을 찾지 못했습니다")
